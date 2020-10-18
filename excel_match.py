import openpyxl
import re
from tqdm import tqdm
from openpyxl import Workbook
from utils.common_functions import get_file_path, get_current_time, get_xls_data, get_xlsx_data

# 实例化
wb_write = Workbook()
# 激活 worksheet
wbs = wb_write.active


def write_merge_excel():
    files = get_file_path('./match', 'excel-match')
    files.sort()
    pbar = tqdm(files)
    data_list = []
    for file_index, file_path in enumerate(pbar):
        pattern_result = re.search(r'/match/(.*)\.xls', file_path)
        colums = pattern_result.group(1).split('-')
        # 匹配的列
        macth_col_index = colums[1]
        # 保留的列
        retain_colums = [int(col_num_str) if col_num_str != '*' else col_num_str for col_num_str in colums[2].split(',')]

        if file_path.endswith('xls'):
            new_rows = get_xls_data(file_path)
        else:
            new_rows = get_xlsx_data(file_path)
        data_map = {}
        for row_values in new_rows:
            new_values = [row_value for row_index, row_value in enumerate(row_values) if
             row_index + 1 in retain_colums or '*' in retain_colums]
            data_map[row_values[int(macth_col_index) - 1]] = [row_value for row_index, row_value in enumerate(row_values) if
                                                         row_index + 1 in retain_colums or '*' in retain_colums]
            data_list.append(data_map)
    print(data_list)

    # 保存文件
    # wb_write.save('./merge/excel-merge' + get_current_time() + '.xlsx')


def format_excel_data(args):
    data_list = []
    # 实例化
    wb_write = Workbook()

    # 激活 worksheet
    wbs = wb_write.active
    for i, file_name in enumerate(get_file_path('./match', 'excel-match')):
        # 载入xlsx文件
        wb = openpyxl.load_workbook(file_name)
        # 循环每个sheet
        for ws_i, ws in enumerate(wb):
            data_map = {}
            for row in ws.rows:
                row_values = [col.value for col in row]
                data_map[row_values[args[i] - 1]] = row_values
            data_list.append(data_map)

    return data_list


def excel_match(*args):
    data_list = format_excel_data(args)
    print(data_list)


write_merge_excel()

# 保存文件
# wb_write.save('excel-merge' + get_current_time() + '.xlsx')
